#!/usr/bin/env python
import argparse
import json
import math
import re
import shutil
import sys
import unicodedata
from datetime import datetime
from pathlib import Path


def normalize_text(value):
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return re.sub(r"\s+", " ", text).strip().lower()


def parse_number(value):
    if value is None:
        return 0
    if isinstance(value, (int, float)) and math.isfinite(value):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0
    text = re.sub(r"[^0-9,.\-]", "", text)
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0


FIELD_ALIASES = {
    "codigo_producto": ["codigo", "cod", "sku", "codigo interno", "codigo producto", "cod producto", "nomenclador"],
    "codigo_barras": ["codigo de barras", "cod barras", "ean", "barcode", "barra"],
    "descripcion": ["descripcion", "producto", "articulo", "mercaderia", "detalle", "nombre"],
    "rubro": ["rubro", "categoria", "category", "familia"],
    "familia": ["familia", "grupo"],
    "marca": ["marca", "brand"],
    "proveedor": ["proveedor", "supplier"],
    "unidad_venta": ["unidad", "unidad venta", "u venta", "presentacion"],
    "costo": ["costo", "cost", "precio costo", "p costo", "costo unitario"],
    "stock": ["stock", "existencia", "cantidad", "stock actual", "stock fisico"],
    "stock_minimo": ["stock minimo", "minimo", "min"],
    "precio_lista_1": ["lista 1", "lista1", "l1", "precio lista 1", "precio_lista_1"],
    "precio_lista_2": ["lista 2", "lista2", "l2", "precio lista 2", "precio_lista_2", "precio"],
    "precio_lista_3": ["lista 3", "lista3", "l3", "precio lista 3", "precio_lista_3"],
    "precio_lista_4": ["lista 4", "lista4", "l4", "precio lista 4", "precio_lista_4"],
    "precio_lista_5": ["lista 5", "lista5", "l5", "precio lista 5", "precio_lista_5"],
    "activo": ["activo", "estado", "habilitado"],
}


def canonical_field(header):
    normalized = normalize_text(header)
    for field, aliases in FIELD_ALIASES.items():
        if normalized in aliases:
            return field
    if not normalized:
        return ""
    if "barra" in normalized or "ean" in normalized or "barcode" in normalized:
        return "codigo_barras"
    if "codigo proveedor" in normalized or "cod proveedor" in normalized:
        return ""
    if normalized.startswith("codigo") or normalized.startswith("cod ") or normalized in {"cod", "sku", "nomenclador"}:
        return "codigo_producto"
    if "descripcion" in normalized or "producto" in normalized or "articulo" in normalized or "mercaderia" in normalized:
        return "descripcion"
    if "subrubro" in normalized:
        return "familia"
    if "categoria" in normalized or normalized == "rubro" or normalized.startswith("rubro "):
        return "rubro"
    if "familia" in normalized or "grupo" in normalized:
        return "familia"
    if "marca" in normalized or "brand" in normalized:
        return "marca"
    if "proveedor" in normalized or "supplier" in normalized:
        return "proveedor"
    if "unid" in normalized or "unidad" in normalized or "presentacion" in normalized:
        return "unidad_venta"
    if "costo" in normalized or normalized == "cost":
        return "costo"
    if "stock minimo" in normalized or normalized == "minimo" or normalized == "min":
        return "stock_minimo"
    if "stock" in normalized or "existencia" in normalized or "cantidad" in normalized:
        return "stock"
    for number in range(1, 6):
        if re.search(rf"(?:lista\s*{number}|lista{number}|\bl{number}\b|precio_lista_{number})", normalized):
            return f"precio_lista_{number}"
    if "activo" in normalized or "estado" in normalized or "habilitado" in normalized:
        return "activo"
    return ""


def load_xlsx_rows(path, sheet_name=""):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("Falta openpyxl en Python. Instalarlo o ejecutar con el runtime bundled de Codex.") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    raw_rows = list(sheet.iter_rows(values_only=True))
    header_index = -1
    fields = []
    for index, row in enumerate(raw_rows[:30]):
        mapped = [canonical_field(cell) for cell in row]
        score = len([field for field in mapped if field])
        if score >= 2 and "descripcion" in mapped:
            header_index = index
            fields = mapped
            break
    if header_index < 0:
        raise RuntimeError("No pude detectar encabezados de productos en el Excel.")
    rows = []
    for row in raw_rows[header_index + 1:]:
        record = {}
        for pos, field in enumerate(fields):
            if field:
                record[field] = row[pos] if pos < len(row) else None
        if any(value not in (None, "") for value in record.values()):
            rows.append(record)
    return rows


def normalize_product(raw, index):
    name = str(raw.get("descripcion") or "").strip()
    if not name:
        raise RuntimeError(f"Fila {index + 1}: producto sin descripcion.")
    code = str(raw.get("codigo_producto") or "").strip() or f"DL-{index + 1:04d}"
    stock = max(0, parse_number(raw.get("stock")))
    cost = max(0, parse_number(raw.get("costo")))
    prices = {f"precio_lista_{i}": max(0, parse_number(raw.get(f"precio_lista_{i}"))) for i in range(1, 6)}
    if not prices["precio_lista_2"]:
        prices["precio_lista_2"] = prices["precio_lista_1"]
    return {
        "codigo_producto": code,
        "codigo_barras": str(raw.get("codigo_barras") or "").strip(),
        "name": name,
        "descripcion": name,
        "rubro": str(raw.get("rubro") or "S/D").strip() or "S/D",
        "familia": str(raw.get("familia") or raw.get("rubro") or "S/D").strip() or "S/D",
        "marca": str(raw.get("marca") or "S/D").strip() or "S/D",
        "proveedor": str(raw.get("proveedor") or "").strip(),
        "unidad_venta": str(raw.get("unidad_venta") or "unidad").strip() or "unidad",
        "stock_fisico": stock,
        "stock_actual": stock,
        "stock": stock,
        "stock_reservado": 0,
        "stock_disponible": stock,
        "stock_en_transito": 0,
        "stock_minimo": max(0, parse_number(raw.get("stock_minimo"))),
        "min": max(0, parse_number(raw.get("stock_minimo"))),
        "costo": cost,
        "cost": cost,
        "margen": 0,
        **prices,
        "price": prices["precio_lista_2"] or prices["precio_lista_1"],
        "priceListId": "PL-L2",
        "priceListName": "Lista Nº 2",
        "activo": "NO" if normalize_text(raw.get("activo")).startswith("inact") else "SI",
        "origen": "importacion-cartera",
        "updatedAt": datetime.utcnow().isoformat() + "Z",
    }


def price_list_item(product, number):
    price = product.get(f"precio_lista_{number}") or product.get("price") or 0
    return {
        "productCode": product.get("codigo_producto", ""),
        "productName": product.get("name", ""),
        "codigo_barras": product.get("codigo_barras", ""),
        "rubro": product.get("rubro", "S/D"),
        "marca": product.get("marca", "S/D"),
        "proveedor": product.get("proveedor", ""),
        "previousPrice": price,
        "price": price,
        "newPrice": price,
        "listNumber": number,
        "difference": 0,
        "percentApplied": 0,
        "marginPct": 0,
        "increasePct": 0,
    }


def build_price_lists(products):
    return [{
        "id": f"PL-L{number}",
        "name": f"Lista Nº {number}",
        "status": "Activa",
        "effectiveAt": "2026-01-01T00:00:00.000Z",
        "isDefault": number == 2,
        "productCount": len(products),
        "updatedAt": datetime.utcnow().isoformat() + "Z",
        "updatedBy": "Importador cartera",
        "rounding": 1,
        "operation": "columna_producto",
        "motive": "Importacion cartera de productos",
        "number": number,
        "generatedFromColumns": True,
        "items": [price_list_item(product, number) for product in products],
    } for number in range(1, 6)]


def backup_files(state_path):
    data_dir = state_path.parent
    stamp = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    backup_dir = data_dir / "backups" / f"{stamp}-importar-cartera-productos"
    backup_dir.mkdir(parents=True, exist_ok=True)
    if state_path.exists():
        shutil.copy2(state_path, backup_dir / state_path.name)
    users_path = data_dir / "users.json"
    if users_path.exists():
        shutil.copy2(users_path, backup_dir / "users.json")
    return backup_dir


def main():
    parser = argparse.ArgumentParser(description="Importa cartera de productos desde Excel al estado del servidor 8790.")
    parser.add_argument("--excel", required=True, help="Ruta del archivo .xlsx")
    parser.add_argument("--state", default=str(Path(__file__).resolve().parents[1] / "data" / "demo-state.json"))
    parser.add_argument("--sheet", default="")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    state_path = Path(args.state)
    if not excel_path.exists():
        raise SystemExit(f"No existe el Excel: {excel_path}")
    payload = json.loads(state_path.read_text(encoding="utf-8"))
    state = payload.get("state") or {}
    rows = load_xlsx_rows(excel_path, args.sheet)
    products = [normalize_product(row, index) for index, row in enumerate(rows)]
    if args.dry_run:
        print(json.dumps({"ok": True, "products": len(products), "sample": products[:5]}, ensure_ascii=False, indent=2))
        return
    backup_dir = backup_files(state_path)
    previous_count = len(state.get("products") or [])
    state["products"] = products
    state["priceLists"] = build_price_lists(products)
    state["priceListAssignments"] = [item for item in state.get("priceListAssignments", []) if normalize_text(item.get("username")) != "kevin"]
    state["priceListAssignments"].insert(0, {
        "username": "kevin",
        "sellerName": "Kevin Guibert",
        "priceListId": "PL-L4",
        "priceListName": "Lista Nº 4",
        "listNumber": 4,
        "locked": True,
        "active": True,
        "updatedAt": datetime.utcnow().isoformat() + "Z",
        "updatedBy": "Importador cartera",
    })
    audit = state.setdefault("productPortfolioAudit", [])
    audit.insert(0, {
        "id": f"PORT-{int(datetime.utcnow().timestamp())}",
        "at": datetime.utcnow().isoformat() + "Z",
        "user": "Importador cartera",
        "username": "script",
        "action": "IMPORTACION_CARTERA_PRODUCTOS",
        "source": str(excel_path),
        "previousCount": previous_count,
        "newCount": len(products),
        "backup": str(backup_dir),
    })
    payload["version"] = int(datetime.utcnow().timestamp() * 1000)
    payload["state"] = state
    state_path.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(json.dumps({"ok": True, "imported": len(products), "previousCount": previous_count, "backup": str(backup_dir)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
